#!/usr/bin/env python3
# -*- coding: UTF-8 -*-
# 2026/4/5 16:40
import os.path

from openpyxl import load_workbook
Desk_path = os.path.join(os.path.expanduser('~'), 'Desktop')
file_path = os.path.join(Desk_path, 'students.xlsx')
wb = load_workbook(file_path)
sheet = wb['考试成绩']
sheet.cell(row=1,column=1).value = '100'
wb.save(file_path)

